"""
Standalone Backup-System für die Installation Business App
Erstellt Backups in verschiedenen Formaten: CSV, Excel, SQLite
"""

import os
import csv
import sqlite3
import zipfile
import tempfile
import shutil
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

# Importiere Models für Datenbankzugriff
from models import db, Customer, Quote, QuoteItem, QuoteSubItem, Order, Invoice, Supplier, SupplierOrder, SupplierOrderItem, PositionTemplate, AcquisitionChannel, CompanySettings, WorkInstruction

class DatabaseBackup:
    def __init__(self):
        self.backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def create_csv_backup(self):
        """Erstellt CSV-Backup aller Tabellen als ZIP-Datei"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Temporary directory für CSV-Dateien
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_files = {}
            
            try:
                # 1. Kunden-Backup
                customers = Customer.query.all()
                if customers:
                    csv_path = os.path.join(temp_dir, 'customers.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Vorname', 'Nachname', 'Email', 'Telefon', 'Adresse', 
                            'PLZ', 'Stadt', 'Status', 'Akquisekanal', 'Termin', 'Kommentare', 'Erstellt'
                        ])
                        for customer in customers:
                            writer.writerow([
                                customer.id,
                                customer.first_name or '',
                                customer.last_name or '',
                                customer.email or '',
                                customer.phone or '',
                                customer.address or '',
                                customer.postal_code or '',
                                customer.city or '',
                                customer.status or '',
                                customer.acquisition_channel.name if customer.acquisition_channel else '',
                                customer.appointment_date.strftime('%d.%m.%Y') if customer.appointment_date else '',
                                customer.comments or '',
                                customer.created_at.strftime('%d.%m.%Y %H:%M') if customer.created_at else ''
                            ])
                    csv_files['customers.csv'] = csv_path
                
                # 2. Angebote-Backup
                quotes = Quote.query.all()
                if quotes:
                    csv_path = os.path.join(temp_dir, 'quotes.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Angebotsnummer', 'Kunde', 'Projekt', 'Status', 
                            'Netto-Betrag', 'MwSt-Betrag', 'Brutto-Betrag', 'Erstellt', 'Gültig bis'
                        ])
                        for quote in quotes:
                            net_total = quote.calculate_net_total() if quote.calculate_net_total() else 0
                            # MwSt berechnen: 20% vom Netto-Betrag
                            vat_amount = net_total * 0.20
                            writer.writerow([
                                quote.id,
                                quote.quote_number or '',
                                quote.customer.full_name if quote.customer else '',
                                quote.project_description or '',
                                quote.status or '',
                                f"{net_total:.2f}",
                                f"{vat_amount:.2f}",
                                f"{quote.total_amount:.2f}" if quote.total_amount else '0.00',
                                quote.created_at.strftime('%d.%m.%Y %H:%M') if quote.created_at else '',
                                quote.valid_until.strftime('%d.%m.%Y') if quote.valid_until else ''
                            ])
                    csv_files['quotes.csv'] = csv_path
                
                # 3. Aufträge-Backup
                orders = Order.query.all()
                if orders:
                    csv_path = os.path.join(temp_dir, 'orders.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Auftragsnummer', 'Angebotsnummer', 'Kunde', 'Status', 
                            'Projektmanager', 'Startdatum', 'Enddatum', 'Erstellt'
                        ])
                        for order in orders:
                            writer.writerow([
                                order.id,
                                order.order_number or '',
                                order.quote.quote_number if order.quote else '',
                                order.quote.customer.full_name if order.quote and order.quote.customer else '',
                                order.status or '',
                                order.project_manager or '',
                                order.start_date.strftime('%d.%m.%Y') if order.start_date else '',
                                order.end_date.strftime('%d.%m.%Y') if order.end_date else '',
                                order.created_at.strftime('%d.%m.%Y %H:%M') if order.created_at else ''
                            ])
                    csv_files['orders.csv'] = csv_path
                
                # 4. Rechnungen-Backup
                invoices = Invoice.query.all()
                if invoices:
                    csv_path = os.path.join(temp_dir, 'invoices.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Rechnungsnummer', 'Auftragsnummer', 'Kunde', 'Typ', 
                            'Prozentsatz', 'Basis-Betrag', 'Netto-Betrag', 'MwSt-Betrag', 'Brutto-Betrag', 
                            'Status', 'Fällig am', 'Bezahlt am', 'Zahlungsreferenz', 'Erstellt', 'Kommentare'
                        ])
                        for invoice in invoices:
                            writer.writerow([
                                invoice.id,
                                invoice.invoice_number or '',
                                invoice.order.order_number if invoice.order else '',
                                invoice.order.quote.customer.full_name if invoice.order and invoice.order.quote and invoice.order.quote.customer else '',
                                invoice.invoice_type or '',
                                f"{invoice.percentage:.1f}%" if invoice.percentage else '',
                                f"{invoice.base_amount:.2f}" if invoice.base_amount else '0.00',
                                f"{invoice.final_amount:.2f}" if invoice.final_amount else '0.00',
                                f"{invoice.vat_amount:.2f}" if invoice.vat_amount else '0.00',
                                f"{invoice.gross_amount:.2f}" if invoice.gross_amount else '0.00',
                                invoice.status or '',
                                invoice.due_date.strftime('%d.%m.%Y') if invoice.due_date else '',
                                invoice.paid_date.strftime('%d.%m.%Y') if invoice.paid_date else '',
                                invoice.payment_reference or '',
                                invoice.created_at.strftime('%d.%m.%Y %H:%M') if invoice.created_at else '',
                                invoice.comments or ''
                            ])
                    csv_files['invoices.csv'] = csv_path
                
                # 5. Lieferanten-Backup
                suppliers = Supplier.query.all()
                if suppliers:
                    csv_path = os.path.join(temp_dir, 'suppliers.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Name', 'Kategorie', 'Kontaktperson', 'Email', 'Telefon', 'Adresse', 'Notizen'
                        ])
                        for supplier in suppliers:
                            writer.writerow([
                                supplier.id,
                                supplier.name or '',
                                supplier.category or '',
                                supplier.contact_person or '',
                                supplier.email or '',
                                supplier.phone or '',
                                supplier.address or '',
                                supplier.notes or ''
                            ])
                    csv_files['suppliers.csv'] = csv_path
                
                # 6. Positionsvorlagen-Backup
                templates = PositionTemplate.query.all()
                if templates:
                    csv_path = os.path.join(temp_dir, 'position_templates.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Name', 'Beschreibung', 'Erstellt'
                        ])
                        for template in templates:
                            writer.writerow([
                                template.id,
                                template.name or '',
                                template.description or '',
                                template.created_at.strftime('%d.%m.%Y %H:%M') if template.created_at else ''
                            ])
                    csv_files['position_templates.csv'] = csv_path
                
                # 7. Firmeneinstellungen-Backup
                settings = CompanySettings.query.all()
                if settings:
                    csv_path = os.path.join(temp_dir, 'company_settings.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Einstellungsname', 'Wert', 'Beschreibung', 'Erstellt', 'Aktualisiert'
                        ])
                        for setting in settings:
                            writer.writerow([
                                setting.id,
                                setting.setting_name or '',
                                setting.setting_value or '',
                                setting.description or '',
                                setting.created_at.strftime('%d.%m.%Y %H:%M') if setting.created_at else '',
                                setting.updated_at.strftime('%d.%m.%Y %H:%M') if setting.updated_at else ''
                            ])
                    csv_files['company_settings.csv'] = csv_path
                
                # 8. Akquisekanäle-Backup
                channels = AcquisitionChannel.query.all()
                if channels:
                    csv_path = os.path.join(temp_dir, 'acquisition_channels.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Name', 'Beschreibung', 'Aktiv', 'Erstellt'
                        ])
                        for channel in channels:
                            writer.writerow([
                                channel.id,
                                channel.name or '',
                                channel.description or '',
                                'Ja' if channel.is_active else 'Nein',
                                channel.created_at.strftime('%d.%m.%Y %H:%M') if channel.created_at else ''
                            ])
                    csv_files['acquisition_channels.csv'] = csv_path
                
                # 9. Lieferantenbestellungen-Backup
                supplier_orders = SupplierOrder.query.all()
                if supplier_orders:
                    csv_path = os.path.join(temp_dir, 'supplier_orders.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Angebotsnummer', 'Auftragsnummer', 'Lieferant', 'Bestelldatum', 
                            'Status', 'Bestätigungsdatum', 'Lieferdatum', 'Notizen'
                        ])
                        for supplier_order in supplier_orders:
                            writer.writerow([
                                supplier_order.id,
                                supplier_order.quote.quote_number if supplier_order.quote else '',
                                supplier_order.order.order_number if supplier_order.order else '',
                                supplier_order.supplier_name or '',
                                supplier_order.order_date.strftime('%d.%m.%Y %H:%M') if supplier_order.order_date else '',
                                supplier_order.status or '',
                                supplier_order.confirmation_date.strftime('%d.%m.%Y %H:%M') if supplier_order.confirmation_date else '',
                                supplier_order.delivery_date.strftime('%d.%m.%Y') if supplier_order.delivery_date else '',
                                supplier_order.notes or ''
                            ])
                    csv_files['supplier_orders.csv'] = csv_path
                
                # 10. Lieferantenbestellungspositionen-Backup
                supplier_order_items = SupplierOrderItem.query.all()
                if supplier_order_items:
                    csv_path = os.path.join(temp_dir, 'supplier_order_items.csv')
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([
                            'ID', 'Bestellungs-ID', 'Unternummer', 'Beschreibung', 'Teilenummer', 'Menge'
                        ])
                        for item in supplier_order_items:
                            writer.writerow([
                                item.id,
                                item.supplier_order_id,
                                item.sub_number or '',
                                item.description or '',
                                item.part_number or '',
                                item.quantity or ''
                            ])
                    csv_files['supplier_order_items.csv'] = csv_path
                
                # 11. Arbeitsanweisungen-Backup (falls vorhanden)
                try:
                    work_instructions = WorkInstruction.query.all()
                    if work_instructions:
                        csv_path = os.path.join(temp_dir, 'work_instructions.csv')
                        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile)
                            writer.writerow([
                                'ID', 'Auftragsnummer', 'Anweisungsnummer', 'Status', 'Erstellt von', 
                                'Arbeitsbeschreibung', 'Besondere Hinweise', 'Geschätzte Dauer', 'Priorität'
                            ])
                            for instruction in work_instructions:
                                writer.writerow([
                                    instruction.id,
                                    instruction.order.order_number if instruction.order else '',
                                    instruction.instruction_number or '',
                                    instruction.status or '',
                                    instruction.created_by or '',
                                    instruction.work_description or '',
                                    instruction.special_instructions or '',
                                    f"{instruction.estimated_duration} Std." if instruction.estimated_duration else '',
                                    instruction.priority or ''
                                ])
                        csv_files['work_instructions.csv'] = csv_path
                except Exception as e:
                    # WorkInstruction Model existiert möglicherweise nicht in allen Versionen
                    print(f"Work Instructions nicht verfügbar: {e}")
                
                # Info-Datei hinzufügen
                info_path = os.path.join(temp_dir, 'BACKUP_INFO.txt')
                with open(info_path, 'w', encoding='utf-8') as info_file:
                    info_content = f"""InnSAN Installation Business App - Datenbank Backup
Erstellt am: {datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')}
Format: CSV (Comma Separated Values)
Encoding: UTF-8
Anzahl Tabellen: {len(csv_files)}

Enthaltene Dateien:
"""
                    for filename in csv_files.keys():
                        info_content += f"- {filename}\n"
                    
                    info_content += """
Hinweise:
- Alle Dateien sind UTF-8 codiert
- Dezimaltrennzeichen: Punkt (.)
- Datumsformat: TT.MM.JJJJ
- Import in Excel: Daten → Aus Text/CSV → UTF-8 auswählen
"""
                    info_file.write(info_content)
                csv_files['BACKUP_INFO.txt'] = info_path
                
            except Exception as e:
                print(f"Fehler beim CSV-Backup: {e}")
                # Leeres ZIP mit Fehlermeldung erstellen
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    zip_file.writestr('ERROR.txt', f'Fehler beim Backup: {str(e)}')
                zip_buffer.seek(0)
                return zip_buffer, f'backup_error_{timestamp}.zip'
            
            # ZIP-Datei erstellen
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Alle Dateien hinzufügen
                for filename, filepath in csv_files.items():
                    zip_file.write(filepath, filename)
            
            zip_buffer.seek(0)
            return zip_buffer, f'InnSAN_Backup_CSV_{timestamp}.zip'
    
    def create_excel_backup(self):
        """Erstellt Excel-Backup aller Tabellen in separaten Sheets"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            # Neue Excel-Arbeitsmappe erstellen
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Standard-Sheet entfernen
            
            # Styling definieren
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="CC5500", end_color="CC5500", fill_type="solid")
            
            # 1. Kunden-Sheet
            customers = Customer.query.all()
            if customers:
                ws = wb.create_sheet("Kunden")
                headers = ['ID', 'Vorname', 'Nachname', 'Email', 'Telefon', 'Adresse', 
                          'PLZ', 'Stadt', 'Status', 'Akquisekanal', 'Termin', 'Kommentare', 'Erstellt']
                ws.append(headers)
                
                # Header formatieren
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Daten hinzufügen
                for customer in customers:
                    ws.append([
                        customer.id,
                        customer.first_name or '',
                        customer.last_name or '',
                        customer.email or '',
                        customer.phone or '',
                        customer.address or '',
                        customer.postal_code or '',
                        customer.city or '',
                        customer.status or '',
                        customer.acquisition_channel.name if customer.acquisition_channel else '',
                        customer.appointment_date if customer.appointment_date else '',
                        customer.comments or '',
                        customer.created_at if customer.created_at else ''
                    ])
            
            # 2. Angebote-Sheet
            quotes = Quote.query.all()
            if quotes:
                ws = wb.create_sheet("Angebote")
                headers = ['ID', 'Angebotsnummer', 'Kunde', 'Projekt', 'Status', 
                          'Netto-Betrag', 'MwSt-Betrag', 'Brutto-Betrag', 'Erstellt', 'Gültig bis']
                ws.append(headers)
                
                # Header formatieren
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for quote in quotes:
                    net_total = quote.calculate_net_total() if quote.calculate_net_total() else 0
                    # MwSt berechnen: 20% vom Netto-Betrag
                    vat_amount = net_total * 0.20
                    ws.append([
                        quote.id,
                        quote.quote_number or '',
                        quote.customer.full_name if quote.customer else '',
                        quote.project_description or '',
                        quote.status or '',
                        net_total,
                        vat_amount,
                        quote.total_amount if quote.total_amount else 0,
                        quote.created_at if quote.created_at else '',
                        quote.valid_until if quote.valid_until else ''
                    ])
            
            # 3. Aufträge-Sheet
            orders = Order.query.all()
            if orders:
                ws = wb.create_sheet("Aufträge")
                headers = ['ID', 'Auftragsnummer', 'Angebotsnummer', 'Kunde', 'Status', 
                          'Projektmanager', 'Startdatum', 'Enddatum', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for order in orders:
                    ws.append([
                        order.id,
                        order.order_number or '',
                        order.quote.quote_number if order.quote else '',
                        order.quote.customer.full_name if order.quote and order.quote.customer else '',
                        order.status or '',
                        order.project_manager or '',
                        order.start_date if order.start_date else '',
                        order.end_date if order.end_date else '',
                        order.created_at if order.created_at else ''
                    ])
            
            # 4. Rechnungen-Sheet
            invoices = Invoice.query.all()
            if invoices:
                ws = wb.create_sheet("Rechnungen")
                headers = ['ID', 'Rechnungsnummer', 'Auftragsnummer', 'Kunde', 'Typ', 
                          'Prozentsatz', 'Basis-Betrag', 'Netto-Betrag', 'MwSt-Betrag', 'Brutto-Betrag', 
                          'Status', 'Fällig am', 'Bezahlt am', 'Zahlungsreferenz', 'Erstellt', 'Kommentare']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for invoice in invoices:
                    ws.append([
                        invoice.id,
                        invoice.invoice_number or '',
                        invoice.order.order_number if invoice.order else '',
                        invoice.order.quote.customer.full_name if invoice.order and invoice.order.quote and invoice.order.quote.customer else '',
                        invoice.invoice_type or '',
                        invoice.percentage if invoice.percentage else 0,
                        invoice.base_amount if invoice.base_amount else 0,
                        invoice.final_amount if invoice.final_amount else 0,
                        invoice.vat_amount if invoice.vat_amount else 0,
                        invoice.gross_amount if invoice.gross_amount else 0,
                        invoice.status or '',
                        invoice.due_date if invoice.due_date else '',
                        invoice.paid_date if invoice.paid_date else '',
                        invoice.payment_reference or '',
                        invoice.created_at if invoice.created_at else '',
                        invoice.comments or ''
                    ])
            
            # 5. Lieferanten-Sheet
            suppliers = Supplier.query.all()
            if suppliers:
                ws = wb.create_sheet("Lieferanten")
                headers = ['ID', 'Name', 'Kategorie', 'Kontaktperson', 'Email', 'Telefon', 'Adresse', 'Notizen']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for supplier in suppliers:
                    ws.append([
                        supplier.id,
                        supplier.name or '',
                        supplier.category or '',
                        supplier.contact_person or '',
                        supplier.email or '',
                        supplier.phone or '',
                        supplier.address or '',
                        supplier.notes or ''
                    ])
            
            # 6. Firmeneinstellungen-Sheet
            settings = CompanySettings.query.all()
            if settings:
                ws = wb.create_sheet("Firmeneinstellungen")
                headers = ['ID', 'Einstellungsname', 'Wert', 'Beschreibung', 'Erstellt', 'Aktualisiert']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for setting in settings:
                    ws.append([
                        setting.id,
                        setting.setting_name or '',
                        setting.setting_value or '',
                        setting.description or '',
                        setting.created_at if setting.created_at else '',
                        setting.updated_at if setting.updated_at else ''
                    ])
            
            # 7. Akquisekanäle-Sheet
            channels = AcquisitionChannel.query.all()
            if channels:
                ws = wb.create_sheet("Akquisekanäle")
                headers = ['ID', 'Name', 'Beschreibung', 'Aktiv', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for channel in channels:
                    ws.append([
                        channel.id,
                        channel.name or '',
                        channel.description or '',
                        channel.is_active,
                        channel.created_at if channel.created_at else ''
                    ])
            
            # 8. Lieferantenbestellungen-Sheet
            supplier_orders = SupplierOrder.query.all()
            if supplier_orders:
                ws = wb.create_sheet("Lieferantenbestellungen")
                headers = ['ID', 'Angebotsnummer', 'Auftragsnummer', 'Lieferant', 'Bestelldatum', 
                          'Status', 'Bestätigungsdatum', 'Lieferdatum', 'Notizen']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for supplier_order in supplier_orders:
                    ws.append([
                        supplier_order.id,
                        supplier_order.quote.quote_number if supplier_order.quote else '',
                        supplier_order.order.order_number if supplier_order.order else '',
                        supplier_order.supplier_name or '',
                        supplier_order.order_date if supplier_order.order_date else '',
                        supplier_order.status or '',
                        supplier_order.confirmation_date if supplier_order.confirmation_date else '',
                        supplier_order.delivery_date if supplier_order.delivery_date else '',
                        supplier_order.notes or ''
                    ])
            
            # 9. Lieferantenbestellungspositionen-Sheet
            supplier_order_items = SupplierOrderItem.query.all()
            if supplier_order_items:
                ws = wb.create_sheet("Bestellpositionen")
                headers = ['ID', 'Bestellungs-ID', 'Unternummer', 'Beschreibung', 'Teilenummer', 'Menge']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for item in supplier_order_items:
                    ws.append([
                        item.id,
                        item.supplier_order_id,
                        item.sub_number or '',
                        item.description or '',
                        item.part_number or '',
                        item.quantity or ''
                    ])
            
            # 10. Arbeitsanweisungen-Sheet (falls vorhanden)
            try:
                work_instructions = WorkInstruction.query.all()
                if work_instructions:
                    ws = wb.create_sheet("Arbeitsanweisungen")
                    headers = ['ID', 'Auftragsnummer', 'Anweisungsnummer', 'Status', 'Erstellt von', 
                              'Arbeitsbeschreibung', 'Besondere Hinweise', 'Geschätzte Dauer', 'Priorität']
                    ws.append(headers)
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    for instruction in work_instructions:
                        ws.append([
                            instruction.id,
                            instruction.order.order_number if instruction.order else '',
                            instruction.instruction_number or '',
                            instruction.status or '',
                            instruction.created_by or '',
                            instruction.work_description or '',
                            instruction.special_instructions or '',
                            instruction.estimated_duration if instruction.estimated_duration else '',
                            instruction.priority or ''
                        ])
            except Exception as e:
                # WorkInstruction Model existiert möglicherweise nicht in allen Versionen
                print(f"Work Instructions nicht verfügbar: {e}")
            
            # Info-Sheet
            ws = wb.create_sheet("Backup-Info")
            ws.append(['InnSAN Installation Business App - Datenbank Backup'])
            ws.append(['Erstellt am:', datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')])
            ws.append(['Format:', 'Microsoft Excel (.xlsx)'])
            ws.append([''])
            ws.append(['Enthaltene Sheets:'])
            for sheet_name in wb.sheetnames:
                if sheet_name != "Backup-Info":
                    ws.append([f'- {sheet_name}'])
            
            # Excel-Datei in BytesIO speichern
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            filename = f'InnSAN_Backup_Excel_{timestamp}.xlsx'
            return excel_buffer, filename
            
        except Exception as e:
            print(f"Fehler beim Excel-Backup: {e}")
            # Fallback: Leere Excel-Datei mit Fehlermeldung
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fehler"
            ws.append(['Fehler beim Backup:', str(e)])
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer, f'backup_error_{timestamp}.xlsx'
    
    def create_sqlite_backup(self):
        """Erstellt SQLite-Backup für lokale Entwicklung (nur bei SQLite-DB)"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            # Check if we're using SQLite (local development)
            from flask import current_app
            database_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
            
            if not database_uri.startswith('sqlite'):
                # Für PostgreSQL: Kein SQLite-Backup möglich
                print("SQLite-Backup nur bei lokaler SQLite-Datenbank verfügbar")
                return None, None
            
            # Suche nach der Datenbank-Datei (nur bei SQLite)
            db_paths = [
                'instance/installation_business.db',
                os.path.join(os.path.dirname(__file__), 'instance', 'installation_business.db'),
                'installation_business.db'
            ]
            
            source_db = None
            for path in db_paths:
                if os.path.exists(path):
                    source_db = path
                    break
            
            if not source_db:
                print("SQLite-Datenbank-Datei nicht gefunden!")
                return None, None
            
            # Datenbank-Datei in BytesIO kopieren
            db_buffer = BytesIO()
            with open(source_db, 'rb') as f:
                db_buffer.write(f.read())
            db_buffer.seek(0)
            
            filename = f'InnSAN_Database_Backup_{timestamp}.db'
            return db_buffer, filename
            
        except Exception as e:
            print(f"Fehler beim SQLite-Backup: {e}")
            return None, None
