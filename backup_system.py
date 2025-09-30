"""
Standalone Backup-System für die Installation Business App
Erstellt Backups in Excel und SQLite Formaten
"""

import os
import sqlite3
import shutil
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

# Importiere Models für Datenbankzugriff
from models import (db, Customer, Quote, QuoteItem, QuoteSubItem, Order, Invoice, 
                   Supplier, SupplierOrder, SupplierOrderItem, PositionTemplate, 
                   AcquisitionChannel, CompanySettings, WorkInstruction, 
                   InvoiceReminder, QuoteRejection, PositionTemplateSubItem,
                   Article, InvoicePosition)
# LoginAdmin bewusst NICHT importiert - Login-Daten bleiben bei Wiederherstellung unverändert

class DatabaseBackup:
    def __init__(self):
        self.backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
        os.makedirs(self.backup_dir, exist_ok=True)
    

    
    def create_excel_backup(self):
        """Erstellt Excel-Backup aller Tabellen in separaten Sheets"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            # Neue Excel-Arbeitsmappe erstellen
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Standard-Sheet entfernen
            
            # Styling definieren
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="CC5500", end_color="CC5500", fill_type="solid")
            
            # 1. Kunden-Sheet
            customers = Customer.query.all()
            if customers:
                ws = wb.create_sheet("Kunden")
                headers = ['ID', 'Anrede', 'Vorname', 'Nachname', 'Email', 'Telefon', 'Adresse', 
                          'PLZ', 'Stadt', 'UID-Nummer', 'Kundenbetreuer', 'Status', 'Akquisekanal', 'Detaillierter Akquisekanal', 
                          '1. Termin', '1. Termin Notizen', '2. Termin', '2. Termin Notizen', 'Kommentare', 'Erstellt']
                ws.append(headers)
                
                # Header formatieren
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Daten hinzufügen
                for customer in customers:
                    ws.append([
                        customer.id,
                        customer.salutation or '',
                        customer.first_name or '',
                        customer.last_name or '',
                        customer.email or '',
                        customer.phone or '',
                        customer.address or '',
                        customer.postal_code or '',
                        customer.city or '',
                        customer.uid_number or '',
                        customer.customer_manager or '',
                        customer.status or '',
                        customer.acquisition_channel.name if customer.acquisition_channel else '',
                        customer.detailed_acquisition_channel or '',
                        customer.appointment_date if customer.appointment_date else '',
                        customer.appointment_notes or '',
                        customer.second_appointment_date if customer.second_appointment_date else '',
                        customer.second_appointment_notes or '',
                        customer.comments or '',
                        customer.created_at if customer.created_at else ''
                    ])
            
            # 2. Angebote-Sheet
            quotes = Quote.query.all()
            if quotes:
                ws = wb.create_sheet("Angebote")
                headers = ['ID', 'Angebotsnummer', 'Kunde', 'Projekt', 'Status', 
                          'Netto-Betrag', 'MwSt-Betrag', 'Brutto-Betrag', 'Erstellt', 'Gültig bis',
                          'Zusatzinfo_anzeigen', 'Preisanzeige_Modus', 'Aufschlag_Prozent',
                          'Leistungsumfang', 'Objektinformationen', 'Installationsleistungen']
                ws.append(headers)
                
                # Header formatieren
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for quote in quotes:
                    net_total = quote.calculate_net_total() if quote.calculate_net_total() else 0
                    # MwSt berechnen: 20% vom Netto-Betrag
                    vat_amount = net_total * 0.20
                    ws.append([
                        quote.id,
                        quote.quote_number or '',
                        quote.customer.full_name if quote.customer else '',
                        quote.project_description or '',
                        quote.status or '',
                        net_total,
                        vat_amount,
                        quote.total_amount if quote.total_amount else 0,
                        quote.created_at if quote.created_at else '',
                        quote.valid_until if quote.valid_until else '',
                        quote.include_additional_info,
                        quote.price_display_mode or 'standard',
                        quote.markup_percentage if quote.markup_percentage else 15.0,
                        quote.leistungsumfang or '',
                        quote.objektinformationen or '',
                        quote.installationsleistungen or ''
                    ])
            
            # 3. Angebotspositionen-Sheet
            quote_items = QuoteItem.query.all()
            if quote_items:
                ws = wb.create_sheet("Angebotspositionen")
                headers = ['ID', 'Angebots_ID', 'Angebotsnummer', 'Positionsnummer', 'Beschreibung', 
                          'Menge', 'Einzelpreis', 'Gesamtpreis', 'Artikeltyp', 'Benötigt_Bestellung', 'Lieferant']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for item in quote_items:
                    ws.append([
                        item.id,
                        item.quote_id,
                        item.quote.quote_number if item.quote else '',
                        item.position_number,
                        item.description or '',
                        item.quantity,
                        item.unit_price,
                        item.total_price,
                        item.item_type or 'standard',
                        item.requires_order,
                        item.supplier or ''
                    ])
            
            # 4. Angebots-Unterartikel-Sheet
            quote_sub_items = QuoteSubItem.query.all()
            if quote_sub_items:
                ws = wb.create_sheet("Angebots_Unterartikel")
                headers = ['ID', 'Position_ID', 'Angebotsnummer', 'Positionsnummer', 'Unternummer', 
                          'Beschreibung', 'Artikeltyp', 'Benötigt_Bestellung', 'Lieferant', 'Teilenummer',
                          'Teil_Menge', 'Teil_Preis', 'Arbeitsstunden', 'Stundensatz', 
                          'Sonstige_Menge', 'Sonstiger_Einheitspreis', 'Berechneter_Preis']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for sub_item in quote_sub_items:
                    ws.append([
                        sub_item.id,
                        sub_item.quote_item_id,
                        sub_item.quote_item.quote.quote_number if sub_item.quote_item and sub_item.quote_item.quote else '',
                        sub_item.quote_item.position_number if sub_item.quote_item else '',
                        sub_item.sub_number or '',
                        sub_item.description or '',
                        sub_item.item_type or 'bestellteil',
                        sub_item.requires_order,
                        sub_item.supplier or '',
                        sub_item.part_number or '',
                        sub_item.part_quantity or '',
                        sub_item.part_price if sub_item.part_price else 0,
                        sub_item.hours if sub_item.hours else 0,
                        sub_item.hourly_rate if sub_item.hourly_rate else 0,
                        sub_item.quantity or '',
                        sub_item.unit_price if sub_item.unit_price else 0,
                        sub_item.price if sub_item.price else 0
                    ])
            
            # 5. Aufträge-Sheet
            orders = Order.query.all()
            if orders:
                ws = wb.create_sheet("Aufträge")
                headers = ['ID', 'Auftragsnummer', 'Angebotsnummer', 'Kunde', 'Status', 
                          'Projektmanager', 'Startdatum', 'Enddatum', 'Notizen', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for order in orders:
                    ws.append([
                        order.id,
                        order.order_number or '',
                        order.quote.quote_number if order.quote else '',
                        order.quote.customer.full_name if order.quote and order.quote.customer else '',
                        order.status or '',
                        order.project_manager or '',
                        order.start_date if order.start_date else '',
                        order.end_date if order.end_date else '',
                        order.notes or '',
                        order.created_at if order.created_at else ''
                    ])
            
            # 6. Rechnungen-Sheet
            invoices = Invoice.query.all()
            if invoices:
                ws = wb.create_sheet("Rechnungen")
                headers = ['ID', 'Rechnungsnummer', 'Auftragsnummer', 'Kunde_ID', 'Kunde_Name', 'Typ', 
                          'Prozentsatz', 'Basis-Betrag', 'Rechnungsbetrag', 'Vorherige_Zahlungen', 'Finaler_Betrag', 
                          'MwSt_Satz', 'MwSt-Betrag', 'Brutto-Betrag', 'Projektname',
                          'Material_Kosten_Edit', 'Arbeitsstunden_Edit', 'Stundensatz_Edit', 'Arbeitskosten_Edit',
                          'Material_Beschreibung', 'Arbeits_Beschreibung', 'Leistungsbeschreibung',
                          'Status', 'Fällig_am', 'Zahlungsziel_Tage', 'Bezahlt_am', 'Zahlungsreferenz',
                          'Bezahlter_Betrag', 'Zahlungskommentar', 'Kommentare', 'Erstellt', 'Aktualisiert']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for invoice in invoices:
                    ws.append([
                        invoice.id,
                        invoice.invoice_number or '',
                        invoice.order.order_number if invoice.order else '',
                        invoice.customer_id or '',
                        (invoice.order.quote.customer.full_name if invoice.order and invoice.order.quote and invoice.order.quote.customer 
                         else invoice.customer.full_name if invoice.customer else ''),
                        invoice.invoice_type or '',
                        invoice.percentage if invoice.percentage else 0,
                        invoice.base_amount if invoice.base_amount else 0,
                        invoice.invoice_amount if invoice.invoice_amount else 0,
                        invoice.previous_payments if invoice.previous_payments else 0,
                        invoice.final_amount if invoice.final_amount else 0,
                        invoice.vat_rate if invoice.vat_rate else 20.0,
                        invoice.vat_amount if invoice.vat_amount else 0,
                        invoice.gross_amount if invoice.gross_amount else 0,
                        invoice.project_name or '',
                        invoice.material_costs_editable if invoice.material_costs_editable else '',
                        invoice.labor_hours_editable if invoice.labor_hours_editable else '',
                        invoice.labor_rate_editable if invoice.labor_rate_editable else '',
                        invoice.labor_costs_editable if invoice.labor_costs_editable else '',
                        invoice.material_description or '',
                        invoice.labor_description or '',
                        invoice.service_description or '',
                        invoice.status or '',
                        invoice.due_date if invoice.due_date else '',
                        invoice.payment_terms if invoice.payment_terms else 14,
                        invoice.paid_date if invoice.paid_date else '',
                        invoice.payment_reference or '',
                        invoice.paid_amount if invoice.paid_amount else 0,
                        invoice.payment_comment or '',
                        invoice.comments or '',
                        invoice.created_at if invoice.created_at else '',
                        invoice.updated_at if invoice.updated_at else ''
                    ])
            
            # 7. Lieferanten-Sheet
            suppliers = Supplier.query.order_by(Supplier.name).all()
            if suppliers:
                ws = wb.create_sheet("Lieferanten")
                headers = ['ID', 'Name', 'Kategorie', 'Kontaktperson', 'Email', 'Telefon', 'Adresse', 'Notizen']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for supplier in suppliers:
                    ws.append([
                        supplier.id,
                        supplier.name or '',
                        supplier.category or '',
                        supplier.contact_person or '',
                        supplier.email or '',
                        supplier.phone or '',
                        supplier.address or '',
                        supplier.notes or ''
                    ])
            
            # 8. Lieferantenbestellungen-Sheet
            supplier_orders = SupplierOrder.query.all()
            if supplier_orders:
                ws = wb.create_sheet("Lieferantenbestellungen")
                headers = ['ID', 'Auftrag_ID', 'Angebot_ID', 'Lieferant', 'Bestelldatum', 'Bestätigungsdatum', 
                          'Liefertermin', 'Status', 'Notizen']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for supplier_order in supplier_orders:
                    ws.append([
                        supplier_order.id,
                        supplier_order.order_id or '',
                        supplier_order.quote_id or '',
                        supplier_order.supplier_name or '',
                        supplier_order.order_date if supplier_order.order_date else '',
                        supplier_order.confirmation_date if supplier_order.confirmation_date else '',
                        supplier_order.delivery_date if supplier_order.delivery_date else '',
                        supplier_order.status or '',
                        supplier_order.notes or '',
                        supplier_order.order_date if supplier_order.order_date else ''
                    ])
            
            # 9. Lieferantenbestellpositionen-Sheet
            supplier_order_items = SupplierOrderItem.query.all()
            if supplier_order_items:
                ws = wb.create_sheet("Lieferantenbestellpositionen")
                headers = ['ID', 'Bestellung_ID', 'Unter_Nr', 'Beschreibung', 'Teilenummer', 
                          'Menge', 'QuoteSubItem_ID']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for item in supplier_order_items:
                    ws.append([
                        item.id,
                        item.supplier_order_id or '',
                        item.sub_number or '',
                        item.description or '',
                        item.part_number or '',
                        item.quantity or 1,
                        item.quote_sub_item_id or ''
                    ])
            
            # 8. Positionsvorlagen-Sheet
            templates = PositionTemplate.query.all()
            if templates:
                ws = wb.create_sheet("Positionsvorlagen")
                headers = ['ID', 'Name', 'Beschreibung', 'Länge_aktiv', 'Breite_aktiv', 'Höhe_aktiv', 
                          'Fläche_aktiv', 'Volumen_aktiv', 'Erstellt', 'Aktualisiert']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for template in templates:
                    ws.append([
                        template.id,
                        template.name or '',
                        template.description or '',
                        template.enable_length,
                        template.enable_width,
                        template.enable_height,
                        template.enable_area,
                        template.enable_volume,
                        template.created_at if template.created_at else '',
                        template.updated_at if template.updated_at else ''
                    ])
            
            # 9. Firmeneinstellungen-Sheet
            settings = CompanySettings.query.all()
            if settings:
                ws = wb.create_sheet("Firmeneinstellungen")
                headers = ['ID', 'Einstellungsname', 'Wert', 'Beschreibung', 'Erstellt', 'Aktualisiert']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for setting in settings:
                    ws.append([
                        setting.id,
                        setting.setting_name or '',
                        setting.setting_value or '',
                        setting.description or '',
                        setting.created_at if setting.created_at else '',
                        setting.updated_at if setting.updated_at else ''
                    ])
            
            # 10. Akquisekanäle-Sheet
            channels = AcquisitionChannel.query.all()
            if channels:
                ws = wb.create_sheet("Akquisekanäle")
                headers = ['ID', 'Name', 'Beschreibung', 'Aktiv', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for channel in channels:
                    ws.append([
                        channel.id,
                        channel.name or '',
                        channel.description or '',
                        channel.is_active,
                        channel.created_at if channel.created_at else ''
                    ])
            
            # 11. Lieferantenbestellungen-Sheet
            supplier_orders = SupplierOrder.query.all()
            if supplier_orders:
                ws = wb.create_sheet("Lieferantenbestellungen")
                headers = ['ID', 'Angebotsnummer', 'Auftragsnummer', 'Lieferant', 'Bestelldatum', 
                          'Status', 'Bestätigungsdatum', 'Lieferdatum', 'Notizen']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for supplier_order in supplier_orders:
                    ws.append([
                        supplier_order.id,
                        supplier_order.quote.quote_number if supplier_order.quote else '',
                        supplier_order.order.order_number if supplier_order.order else '',
                        supplier_order.supplier_name or '',
                        supplier_order.order_date if supplier_order.order_date else '',
                        supplier_order.status or '',
                        supplier_order.confirmation_date if supplier_order.confirmation_date else '',
                        supplier_order.delivery_date if supplier_order.delivery_date else '',
                        supplier_order.notes or ''
                    ])
            
            # 12. Lieferantenbestellungspositionen-Sheet
            supplier_order_items = SupplierOrderItem.query.all()
            if supplier_order_items:
                ws = wb.create_sheet("Bestellpositionen")
                headers = ['ID', 'Bestellungs-ID', 'Unternummer', 'Beschreibung', 'Teilenummer', 'Menge', 'Angebots-Unterposition-ID']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for item in supplier_order_items:
                    ws.append([
                        item.id,
                        item.supplier_order_id,
                        item.sub_number or '',
                        item.description or '',
                        item.part_number or '',
                        item.quantity or '',
                        item.quote_sub_item_id or ''
                    ])
            
            # 13. Arbeitsanweisungen-Sheet (falls vorhanden)
            try:
                work_instructions = WorkInstruction.query.all()
                if work_instructions:
                    ws = wb.create_sheet("Arbeitsanweisungen")
                    headers = ['ID', 'Auftragsnummer', 'Anweisungsnummer', 'Status', 'Erstellt von', 
                              'Sonstiges', 'Benötigte_Werkzeuge', 'Geschätzte_Dauer', 'Priorität',
                              'Montageort', 'Zugangserfordernisse', 'PDF_Pfad', 'Hat_Fotos', 'Hat_3D_Plan',
                              'Foto_Pfade', 'Plan_Pfad', 'Tatsächlicher_Start', 'Tatsächliches_Ende',
                              'Tatsächliche_Dauer', 'Abschluss_Notizen', 'Qualitätskontrolle', 'Erstellt', 'Aktualisiert']
                    ws.append(headers)
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    for instruction in work_instructions:
                        ws.append([
                            instruction.id,
                            instruction.order.order_number if instruction.order else '',
                            instruction.instruction_number or '',
                            instruction.status or '',
                            instruction.created_by or '',
                            instruction.sonstiges or '',
                            instruction.tools_required or '',
                            instruction.estimated_duration if instruction.estimated_duration else '',
                            instruction.priority or '',
                            instruction.installation_location or '',
                            instruction.access_requirements or '',
                            instruction.pdf_path or '',
                            instruction.has_photos,
                            instruction.has_3d_plan,
                            instruction.photo_paths or '',
                            instruction.plan_path or '',
                            instruction.actual_start_time if instruction.actual_start_time else '',
                            instruction.actual_end_time if instruction.actual_end_time else '',
                            instruction.actual_duration if instruction.actual_duration else '',
                            instruction.completion_notes or '',
                            instruction.quality_check,
                            instruction.created_at if instruction.created_at else '',
                            instruction.updated_at if instruction.updated_at else ''
                        ])
            except Exception as e:
                # WorkInstruction Model existiert möglicherweise nicht in allen Versionen
                print(f"Work Instructions nicht verfügbar: {e}")
            
            # 14. Rechnungs-Reminder-Sheet
            reminders = InvoiceReminder.query.all()
            if reminders:
                ws = wb.create_sheet("Rechnungs_Reminder")
                headers = ['ID', 'Auftrag_ID', 'Auftragsnummer', 'Kunde', 'Reminder_Typ', 
                          'Fällig_am', 'Ist_Ausgeblendet', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for reminder in reminders:
                    ws.append([
                        reminder.id,
                        reminder.order_id,
                        reminder.order.order_number if reminder.order else '',
                        reminder.order.quote.customer.full_name if reminder.order and reminder.order.quote and reminder.order.quote.customer else '',
                        reminder.reminder_type or '',
                        reminder.due_date if reminder.due_date else '',
                        reminder.is_dismissed,
                        reminder.created_at if reminder.created_at else ''
                    ])
            
            # 15. Angebots-Ablehnungen-Sheet
            quote_rejections = QuoteRejection.query.all()
            if quote_rejections:
                ws = wb.create_sheet("Angebots_Ablehnungen")
                headers = ['ID', 'Angebots_ID', 'Angebotsnummer', 'Kunde', 'Ablehnungsgrund', 'Erstellt']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for rejection in quote_rejections:
                    ws.append([
                        rejection.id,
                        rejection.quote_id,
                        rejection.quote.quote_number if rejection.quote else '',
                        rejection.quote.customer.full_name if rejection.quote and rejection.quote.customer else '',
                        rejection.reason or '',
                        rejection.created_at if rejection.created_at else ''
                    ])
            
            # 16. Positionsvorlagen-Unterartikel-Sheet
            template_subitems = PositionTemplateSubItem.query.all()
            if template_subitems:
                ws = wb.create_sheet("Vorlagen_Unterartikel")
                headers = ['ID', 'Positionsvorlage_ID', 'Vorlage_Name', 'Beschreibung', 'Artikeltyp', 
                          'Einheit', 'Preis_pro_Einheit', 'Formel', 'Position', 
                          'Benötigt_Bestellung', 'Lieferant', 'Teilenummer', 'Teil_Menge', 'Teil_Preis',
                          'Arbeitsstunden', 'Stundensatz', 'Menge', 'Einheitspreis', 'Berechneter_Preis']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for subitem in template_subitems:
                    ws.append([
                        subitem.id,
                        subitem.template_id,
                        subitem.template.name if subitem.template else '',
                        subitem.description or '',
                        subitem.item_type or '',
                        subitem.unit or '',
                        subitem.price_per_unit if subitem.price_per_unit else 0,
                        subitem.formula or '',
                        subitem.position if subitem.position else 0,
                        subitem.requires_order,
                        subitem.supplier or '',
                        subitem.part_number or '',
                        subitem.part_quantity or '',
                        subitem.part_price if subitem.part_price else 0,
                        subitem.hours if subitem.hours else 0,
                        subitem.hourly_rate if subitem.hourly_rate else 0,
                        subitem.quantity or '',
                        subitem.unit_price if subitem.unit_price else 0,
                        subitem.price if subitem.price else 0
                    ])
            
            # 17. Admin-Benutzer werden aus Sicherheitsgründen NICHT im Backup gespeichert
            # Login-Daten bleiben bei Wiederherstellung unverändert
            
            # 18. Artikel-Sheet
            articles = Article.query.all()
            if articles:
                ws = wb.create_sheet("Artikel")
                headers = ['ID', 'Name', 'Beschreibung']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for article in articles:
                    ws.append([
                        article.id,
                        article.name or '',
                        article.description or ''
                    ])
            
            # 19. Rechnungspositionen-Sheet
            invoice_positions = InvoicePosition.query.all()
            if invoice_positions:
                ws = wb.create_sheet("Rechnungspositionen")
                headers = ['ID', 'Rechnung_ID', 'Position_Nr', 'Artikel_ID', 'Artikel_Text', 'Beschreibung', 
                          'Menge', 'Einheit', 'Preis_Netto', 'Preis_Brutto', 'Rabatt_Wert', 'Rabatt_Typ', 
                          'Zeilensumme_Netto', 'Zeilensumme_Brutto', 'MwSt_Satz']
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for position in invoice_positions:
                    ws.append([
                        position.id,
                        position.invoice_id,
                        position.position_number,
                        position.article_id or '',
                        position.article_text or '',
                        position.description or '',
                        position.quantity if position.quantity else 0,
                        position.unit or '',
                        position.price_net if position.price_net else 0,
                        position.price_gross if position.price_gross else 0,
                        position.discount_value if position.discount_value else 0,
                        position.discount_type or '',
                        position.line_total_net if position.line_total_net else 0,
                        position.line_total_gross if position.line_total_gross else 0,
                        position.vat_rate if position.vat_rate else 0
                    ])
            
            # Info-Sheet
            ws = wb.create_sheet("Backup-Info")
            ws.append(['InnSAN Installation Business App - Datenbank Backup'])
            ws.append(['Erstellt am:', datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')])
            ws.append(['Format:', 'Microsoft Excel (.xlsx)'])
            ws.append([''])
            ws.append(['Enthaltene Sheets:'])
            for sheet_name in wb.sheetnames:
                if sheet_name != "Backup-Info":
                    ws.append([f'- {sheet_name}'])
            
            # Excel-Datei in BytesIO speichern
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            filename = f'InnSAN_Backup_Excel_{timestamp}.xlsx'
            return excel_buffer, filename
            
        except Exception as e:
            print(f"Fehler beim Excel-Backup: {e}")
            # Fallback: Leere Excel-Datei mit Fehlermeldung
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fehler"
            ws.append(['Fehler beim Backup:', str(e)])
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer, f'backup_error_{timestamp}.xlsx'
    
    def save_backup_to_disk(self, buffer, filename):
        """Speichert ein Backup-Buffer als physische Datei im backups-Ordner"""
        try:
            file_path = os.path.join(self.backup_dir, filename)
            with open(file_path, 'wb') as f:
                f.write(buffer.getvalue())
            print(f"✓ Backup gespeichert: {file_path}")
            return file_path
        except Exception as e:
            print(f"Fehler beim Speichern der Backup-Datei: {e}")
            return None

    def create_complete_backup(self):
        """Erstellt Excel- und SQLite-Backup und speichert beide als Dateien"""
        print("Erstelle vollständiges Backup...")
        
        # Excel-Backup erstellen und speichern
        excel_buffer, excel_filename = self.create_excel_backup()
        if excel_buffer:
            excel_path = self.save_backup_to_disk(excel_buffer, excel_filename)
        
        # SQLite-Backup erstellen und speichern
        sqlite_buffer, sqlite_filename = self.create_sqlite_backup()
        if sqlite_buffer:
            sqlite_path = self.save_backup_to_disk(sqlite_buffer, sqlite_filename)
        
        return {
            'excel': {'buffer': excel_buffer, 'filename': excel_filename, 'path': excel_path if excel_buffer else None},
            'sqlite': {'buffer': sqlite_buffer, 'filename': sqlite_filename, 'path': sqlite_path if sqlite_buffer else None}
        }

    def create_sqlite_backup(self):
        """Erstellt eine SQLite-Datenbank - entweder als Kopie (lokal) oder Export (PostgreSQL)"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            # Prüfe ob wir eine lokale SQLite-Datei haben (lokale Entwicklung)
            db_paths = [
                'instance/installation_business.db',
                os.path.join(os.path.dirname(__file__), 'instance', 'installation_business.db'),
                'installation_business.db'
            ]
            
            source_db = None
            for path in db_paths:
                if os.path.exists(path):
                    source_db = path
                    break
            
            if source_db:
                # Lokale SQLite-Datei - erstelle gefilterte Kopie OHNE login_admins
                print("Lokale SQLite erkannt - erstelle gefilterte Kopie ohne Login-Daten...")
                db_buffer = self._create_filtered_sqlite_copy(source_db)
                if not db_buffer:
                    print("Fehler beim Erstellen der gefilterten Kopie!")
                    return None, None
            else:
                # PostgreSQL (Railway) - neue SQLite-Datei aus den Daten erstellen
                print("PostgreSQL erkannt - erstelle SQLite-Export...")
                db_buffer = self._create_sqlite_from_postgresql()
                if not db_buffer:
                    print("Fehler beim PostgreSQL-Export!")
                    return None, None
            
            filename = f'InnSAN_Database_Backup_{timestamp}.db'
            return db_buffer, filename
            
        except Exception as e:
            print(f"Fehler beim SQLite-Backup: {e}")
            return None, None
    
    def _create_filtered_sqlite_copy(self, source_db_path):
        """Erstellt eine gefilterte SQLite-Kopie OHNE login_admins Tabelle"""
        try:
            # Temporäre SQLite-Datei erstellen
            db_buffer = BytesIO()
            
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as temp_file:
                temp_db_path = temp_file.name
            
            # Neue SQLite-Verbindung für gefilterte Kopie
            dest_conn = sqlite3.connect(temp_db_path)
            dest_cursor = dest_conn.cursor()
            
            # Original-DB öffnen
            source_conn = sqlite3.connect(source_db_path)
            source_cursor = source_conn.cursor()
            
            # Alle Tabellen AUSSER login_admins kopieren
            source_cursor.execute("SELECT name, sql FROM sqlite_master WHERE type='table'")
            tables = source_cursor.fetchall()
            
            excluded_tables = ['login_admins']  # Admin-Tabellen ausschließen
            
            for table_name, create_sql in tables:
                if table_name not in excluded_tables and not table_name.startswith('sqlite_'):
                    # Tabelle erstellen
                    dest_cursor.execute(create_sql)
                    
                    # Daten kopieren
                    source_cursor.execute(f'SELECT * FROM `{table_name}`')
                    rows = source_cursor.fetchall()
                    
                    if rows:
                        # Column count ermitteln
                        source_cursor.execute(f'PRAGMA table_info(`{table_name}`)')
                        columns = source_cursor.fetchall()
                        placeholders = ','.join(['?' for _ in columns])
                        
                        dest_cursor.executemany(f'INSERT INTO `{table_name}` VALUES ({placeholders})', rows)
            
            source_conn.close()
            dest_conn.commit()
            dest_conn.close()
            
            # Datei in BytesIO laden
            with open(temp_db_path, 'rb') as f:
                db_buffer.write(f.read())
            db_buffer.seek(0)
            
            # Temporäre Datei löschen
            os.unlink(temp_db_path)
            
            print("✅ Gefilterte SQLite-Kopie erstellt (ohne login_admins)")
            return db_buffer
            
        except Exception as e:
            print(f"Fehler beim Erstellen der gefilterten Kopie: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _create_sqlite_from_postgresql(self):
        """Erstellt eine SQLite-Datei aus PostgreSQL-Daten"""
        try:
            # Temporäre SQLite-Datei in BytesIO erstellen
            db_buffer = BytesIO()
            
            # Temporäre SQLite-Verbindung erstellen
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as temp_file:
                temp_db_path = temp_file.name
            
            # SQLite-Verbindung öffnen und Schema erstellen
            sqlite_conn = sqlite3.connect(temp_db_path)
            cursor = sqlite_conn.cursor()
            
            # Schema für alle wichtigen Tabellen erstellen
            self._create_sqlite_schema(cursor)
            
            # Daten aus PostgreSQL lesen und in SQLite einfügen
            self._export_data_to_sqlite(cursor)
            
            sqlite_conn.commit()
            sqlite_conn.close()
            
            # Datei in BytesIO laden
            with open(temp_db_path, 'rb') as f:
                db_buffer.write(f.read())
            db_buffer.seek(0)
            
            # Temporäre Datei löschen
            os.unlink(temp_db_path)
            
            return db_buffer
            
        except Exception as e:
            print(f"Fehler beim PostgreSQL-zu-SQLite Export: {e}")
            return None
    
    def _create_sqlite_schema(self, cursor):
        """Erstellt das SQLite-Schema basierend auf den Flask-SQLAlchemy Models"""
        # Vereinfachtes Schema - die wichtigsten Tabellen
        schema_sql = """
        CREATE TABLE customer (
            id INTEGER PRIMARY KEY,
            salutation VARCHAR(100),
            first_name VARCHAR(100) NOT NULL,
            last_name VARCHAR(100) NOT NULL,
            email VARCHAR(120) NOT NULL,
            phone VARCHAR(20),
            address TEXT,
            city VARCHAR(100),
            postal_code VARCHAR(10),
            uid_number VARCHAR(20),
            customer_manager VARCHAR(100),
            acquisition_channel_id INTEGER,
            detailed_acquisition_channel TEXT,
            status VARCHAR(50) DEFAULT '1. Termin vereinbaren',
            appointment_date DATE,
            appointment_notes TEXT,
            second_appointment_date DATE,
            second_appointment_notes TEXT,
            comments TEXT,
            created_at DATETIME
        );
        
        CREATE TABLE quote (
            id INTEGER PRIMARY KEY,
            quote_number VARCHAR(50) UNIQUE NOT NULL,
            customer_id INTEGER NOT NULL,
            project_description TEXT,
            total_amount FLOAT DEFAULT 0.0,
            status VARCHAR(50) DEFAULT 'Entwurf',
            created_at DATETIME,
            valid_until DATE,
            include_additional_info BOOLEAN DEFAULT 1,
            price_display_mode VARCHAR(20) DEFAULT 'standard',
            show_subitem_prices BOOLEAN DEFAULT 0,
            markup_percentage FLOAT DEFAULT 15.0,
            leistungsumfang TEXT,
            objektinformationen TEXT,
            installationsleistungen TEXT
        );
        
        CREATE TABLE quote_item (
            id INTEGER PRIMARY KEY,
            quote_id INTEGER NOT NULL,
            quantity FLOAT NOT NULL DEFAULT 1.0,
            unit_price FLOAT NOT NULL DEFAULT 0.0,
            total_price FLOAT NOT NULL DEFAULT 0.0,
            description TEXT NOT NULL,
            position_number INTEGER NOT NULL DEFAULT 1,
            requires_order BOOLEAN DEFAULT 0,
            supplier VARCHAR(200),
            item_type VARCHAR(20) DEFAULT 'standard'
        );
        
        CREATE TABLE quote_sub_item (
            id INTEGER PRIMARY KEY,
            quote_item_id INTEGER NOT NULL,
            sub_number VARCHAR(10) NOT NULL,
            description TEXT NOT NULL,
            item_type VARCHAR(20) NOT NULL DEFAULT 'bestellteil',
            requires_order BOOLEAN DEFAULT 0,
            supplier VARCHAR(200),
            part_number VARCHAR(100),
            part_quantity VARCHAR(50) DEFAULT '1',
            part_price FLOAT DEFAULT 0.0,
            hours FLOAT DEFAULT 0.0,
            hourly_rate FLOAT DEFAULT 95.0,
            quantity VARCHAR(50) DEFAULT '',
            unit_price FLOAT DEFAULT 0.0,
            price FLOAT DEFAULT 0.0
        );
        
        CREATE TABLE position_templates (
            id INTEGER PRIMARY KEY,
            name VARCHAR(128) NOT NULL,
            description TEXT,
            created_at DATETIME,
            updated_at DATETIME,
            enable_length BOOLEAN DEFAULT 0,
            enable_width BOOLEAN DEFAULT 0,
            enable_height BOOLEAN DEFAULT 0,
            enable_area BOOLEAN DEFAULT 0,
            enable_volume BOOLEAN DEFAULT 0
        );
        
        CREATE TABLE position_template_subitems (
            id INTEGER PRIMARY KEY,
            template_id INTEGER NOT NULL,
            description VARCHAR(256) NOT NULL,
            item_type VARCHAR(32) NOT NULL,
            unit VARCHAR(32),
            price_per_unit FLOAT,
            formula VARCHAR(128),
            position INTEGER DEFAULT 0,
            requires_order BOOLEAN DEFAULT 0,
            supplier VARCHAR(200),
            part_number VARCHAR(100),
            part_quantity VARCHAR(50) DEFAULT '1',
            part_price FLOAT DEFAULT 0.0,
            hours FLOAT DEFAULT 0.0,
            hourly_rate FLOAT DEFAULT 95.0,
            quantity VARCHAR(50) DEFAULT '',
            unit_price FLOAT DEFAULT 0.0,
            price FLOAT DEFAULT 0.0
        );
        
        CREATE TABLE invoice (
            id INTEGER PRIMARY KEY,
            invoice_number VARCHAR(50) UNIQUE NOT NULL,
            order_id INTEGER,
            customer_id INTEGER,
            invoice_type VARCHAR(20) NOT NULL,
            percentage FLOAT NOT NULL,
            base_amount FLOAT NOT NULL,
            invoice_amount FLOAT NOT NULL,
            previous_payments FLOAT DEFAULT 0.0,
            final_amount FLOAT NOT NULL,
            vat_rate FLOAT DEFAULT 20.0,
            vat_amount FLOAT NOT NULL,
            gross_amount FLOAT NOT NULL,
            project_name VARCHAR(255),
            due_date DATE NOT NULL,
            payment_terms INTEGER DEFAULT 14,
            status VARCHAR(20) DEFAULT 'erstellt',
            paid_date DATE,
            payment_reference VARCHAR(100),
            comments TEXT,
            service_description TEXT,
            created_at DATETIME,
            updated_at DATETIME,
            paid_amount FLOAT DEFAULT 0.0,
            payment_comment TEXT
        );
        
        -- login_admins Tabelle wird bewusst NICHT erstellt
        -- Login-Daten bleiben bei Wiederherstellung unverändert
        
        CREATE TABLE supplier_order (
            id INTEGER PRIMARY KEY,
            order_id INTEGER,
            quote_id INTEGER,
            supplier_name VARCHAR(255) NOT NULL,
            order_date DATE NOT NULL,
            confirmation_date DATE,
            delivery_date DATE,
            status VARCHAR(50) DEFAULT 'Noch nicht bestellt',
            notes TEXT
        );
        
        CREATE TABLE supplier_order_item (
            id INTEGER PRIMARY KEY,
            supplier_order_id INTEGER NOT NULL,
            sub_number VARCHAR(10) NOT NULL,
            description TEXT NOT NULL,
            part_number VARCHAR(100),
            quantity VARCHAR(50) DEFAULT '1',
            quote_sub_item_id INTEGER,
            FOREIGN KEY (supplier_order_id) REFERENCES supplier_order (id) ON DELETE CASCADE,
            FOREIGN KEY (quote_sub_item_id) REFERENCES quote_sub_item (id)
        );
        
        CREATE TABLE article (
            id INTEGER PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            description TEXT
        );
        
        CREATE TABLE invoice_position (
            id INTEGER PRIMARY KEY,
            invoice_id INTEGER NOT NULL,
            position_number INTEGER NOT NULL,
            article_id INTEGER,
            article_text VARCHAR(200),
            description TEXT,
            quantity DECIMAL(10,3) DEFAULT 1.0,
            unit VARCHAR(20) DEFAULT 'Stk',
            price_net DECIMAL(10,2) NOT NULL,
            price_gross DECIMAL(10,2) NOT NULL,
            discount_value DECIMAL(10,2) DEFAULT 0.0,
            discount_type VARCHAR(10) DEFAULT '€',
            line_total_net DECIMAL(10,2),
            line_total_gross DECIMAL(10,2),
            vat_rate DECIMAL(5,2) DEFAULT 20.0,
            FOREIGN KEY (invoice_id) REFERENCES invoice (id) ON DELETE CASCADE,
            FOREIGN KEY (article_id) REFERENCES article (id) ON DELETE SET NULL
        );
        """
        
        # Schema ausführen
        for statement in schema_sql.split(';'):
            if statement.strip():
                cursor.execute(statement.strip())
    
    def _export_data_to_sqlite(self, cursor):
        """Exportiert Daten von PostgreSQL zu SQLite"""
        try:
            # Kunden exportieren
            customers = Customer.query.all()
            for customer in customers:
                cursor.execute('''
                    INSERT INTO customer (id, salutation, first_name, last_name, email, phone, 
                                        address, city, postal_code, uid_number, customer_manager, 
                                        acquisition_channel_id, detailed_acquisition_channel,
                                        status, appointment_date, appointment_notes, 
                                        second_appointment_date, second_appointment_notes, 
                                        comments, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    customer.id, customer.salutation, customer.first_name, customer.last_name,
                    customer.email, customer.phone, customer.address, customer.city,
                    customer.postal_code, customer.uid_number, customer.customer_manager, customer.acquisition_channel_id,
                    customer.detailed_acquisition_channel, customer.status, customer.appointment_date,
                    customer.appointment_notes, customer.second_appointment_date, 
                    customer.second_appointment_notes, customer.comments, customer.created_at
                ))
            
            # Angebote exportieren
            quotes = Quote.query.all()
            for quote in quotes:
                cursor.execute('''
                    INSERT INTO quote (id, quote_number, customer_id, project_description, 
                                     total_amount, status, created_at, valid_until, 
                                     include_additional_info, price_display_mode, 
                                     show_subitem_prices, markup_percentage, leistungsumfang,
                                     objektinformationen, installationsleistungen)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    quote.id, quote.quote_number, quote.customer_id, quote.project_description,
                    quote.total_amount, quote.status, quote.created_at, quote.valid_until,
                    quote.include_additional_info, quote.price_display_mode,
                    quote.show_subitem_prices, quote.markup_percentage, quote.leistungsumfang,
                    quote.objektinformationen, quote.installationsleistungen
                ))
            
            # Angebotspositionen exportieren
            quote_items = QuoteItem.query.all()
            for item in quote_items:
                cursor.execute('''
                    INSERT INTO quote_item (id, quote_id, quantity, unit_price, total_price,
                                          description, position_number, requires_order, 
                                          supplier, item_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    item.id, item.quote_id, item.quantity, item.unit_price, item.total_price,
                    item.description, item.position_number, item.requires_order,
                    item.supplier, item.item_type
                ))
            
            # Angebots-Unterpositionen exportieren
            quote_sub_items = QuoteSubItem.query.all()
            for sub_item in quote_sub_items:
                cursor.execute('''
                    INSERT INTO quote_sub_item (id, quote_item_id, sub_number, description, 
                                              item_type, requires_order, supplier, part_number,
                                              part_quantity, part_price, hours, hourly_rate,
                                              quantity, unit_price, price)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    sub_item.id, sub_item.quote_item_id, sub_item.sub_number, sub_item.description,
                    sub_item.item_type, sub_item.requires_order, sub_item.supplier, sub_item.part_number,
                    sub_item.part_quantity, sub_item.part_price, sub_item.hours, sub_item.hourly_rate,
                    sub_item.quantity, sub_item.unit_price, sub_item.price
                ))
            
            # Positionsvorlagen exportieren
            templates = PositionTemplate.query.all()
            for template in templates:
                cursor.execute('''
                    INSERT INTO position_templates (id, name, description, created_at, updated_at,
                                                  enable_length, enable_width, enable_height,
                                                  enable_area, enable_volume)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    template.id, template.name, template.description, template.created_at, 
                    template.updated_at, template.enable_length, template.enable_width,
                    template.enable_height, template.enable_area, template.enable_volume
                ))
            
            # Positionsvorlagen-Unterpositionen exportieren
            template_subitems = PositionTemplateSubItem.query.all()
            for subitem in template_subitems:
                cursor.execute('''
                    INSERT INTO position_template_subitems (id, template_id, description, item_type,
                                                          unit, price_per_unit, formula, position, 
                                                          requires_order, supplier, part_number,
                                                          part_quantity, part_price, hours, hourly_rate,
                                                          quantity, unit_price, price)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    subitem.id, subitem.template_id, subitem.description, subitem.item_type,
                    subitem.unit, subitem.price_per_unit, subitem.formula, subitem.position,
                    subitem.requires_order, subitem.supplier, subitem.part_number,
                    subitem.part_quantity, subitem.part_price, subitem.hours, subitem.hourly_rate,
                    subitem.quantity, subitem.unit_price, subitem.price
                ))
            
            # Rechnungen exportieren
            invoices = Invoice.query.all()
            for invoice in invoices:
                # Basis-Felder die sicher existieren
                cursor.execute('''
                    INSERT INTO invoice (id, invoice_number, order_id, customer_id, invoice_type,
                                       percentage, base_amount, invoice_amount, previous_payments,
                                       final_amount, vat_rate, vat_amount, gross_amount,
                                       project_name, due_date, payment_terms, status, paid_date,
                                       payment_reference, comments, service_description, created_at, 
                                       updated_at, paid_amount, payment_comment)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    invoice.id, invoice.invoice_number, invoice.order_id, invoice.customer_id,
                    invoice.invoice_type, invoice.percentage, invoice.base_amount, 
                    invoice.invoice_amount, invoice.previous_payments, invoice.final_amount,
                    invoice.vat_rate, invoice.vat_amount, invoice.gross_amount, invoice.project_name,
                    invoice.due_date, invoice.payment_terms, invoice.status, invoice.paid_date,
                    invoice.payment_reference, invoice.comments, invoice.service_description, 
                    invoice.created_at, invoice.updated_at, 
                    getattr(invoice, 'paid_amount', 0.0),
                    getattr(invoice, 'payment_comment', None)
                ))
            
            # Lieferantenbestellungen exportieren
            supplier_orders = SupplierOrder.query.all()
            for supplier_order in supplier_orders:
                cursor.execute('''
                    INSERT INTO supplier_order (id, order_id, quote_id, supplier_name, order_date, 
                                              confirmation_date, delivery_date, status, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    supplier_order.id, supplier_order.order_id, supplier_order.quote_id,
                    supplier_order.supplier_name, supplier_order.order_date,
                    supplier_order.confirmation_date, supplier_order.delivery_date,
                    supplier_order.status, supplier_order.notes
                ))            # Lieferantenbestellpositionen exportieren
            supplier_order_items = SupplierOrderItem.query.all()
            for item in supplier_order_items:
                cursor.execute('''
                    INSERT INTO supplier_order_item (id, supplier_order_id, sub_number, 
                                                    description, part_number, quantity, 
                                                    quote_sub_item_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    item.id, item.supplier_order_id, item.sub_number, item.description,
                    item.part_number, item.quantity, item.quote_sub_item_id
                ))
            
            # Artikel exportieren
            articles = Article.query.all()
            for article in articles:
                cursor.execute('''
                    INSERT INTO article (id, name, description)
                    VALUES (?, ?, ?)
                ''', (
                    article.id, article.name, article.description
                ))
            
            # Rechnungspositionen exportieren
            invoice_positions = InvoicePosition.query.all()
            for position in invoice_positions:
                cursor.execute('''
                    INSERT INTO invoice_position (id, invoice_id, position_number, article_id, 
                                                article_text, description, quantity, unit, 
                                                price_net, price_gross, discount_value, 
                                                discount_type, line_total_net, line_total_gross, 
                                                vat_rate)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    position.id, position.invoice_id, position.position_number, position.article_id,
                    position.article_text, position.description, 
                    float(position.quantity) if position.quantity else 0.0, position.unit,
                    float(position.price_net) if position.price_net else 0.0, 
                    float(position.price_gross) if position.price_gross else 0.0, 
                    float(position.discount_value) if position.discount_value else 0.0, 
                    position.discount_type, 
                    float(position.line_total_net) if position.line_total_net else 0.0, 
                    float(position.line_total_gross) if position.line_total_gross else 0.0,
                    float(position.vat_rate) if position.vat_rate else 20.0
                ))
            
            print(f"SQLite-Export erfolgreich: {len(customers)} Kunden, {len(quotes)} Angebote, {len(quote_items)} Positionen, {len(quote_sub_items)} Unterpositionen, {len(templates)} Vorlagen, {len(template_subitems)} Vorlagen-Unterpositionen, {len(invoices)} Rechnungen, {len(articles)} Artikel, {len(invoice_positions)} Rechnungspositionen, {len(supplier_orders)} Lieferantenbestellungen, {len(supplier_order_items)} Bestellpositionen")
            
        except Exception as e:
            print(f"Fehler beim Datenexport: {e}")
            raise e
